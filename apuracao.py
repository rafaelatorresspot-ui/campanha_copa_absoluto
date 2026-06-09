"""
╔══════════════════════════════════════════════════════════════════╗
║         COPA BRACELL ABSOLUTO — APURAÇÃO AUTOMÁTICA VIA SQL      ║
╚══════════════════════════════════════════════════════════════════╝

Fluxo:
  SQL Server  +  metas_manuais.json
    → apuracao_sql.py  (este script)
      → controle_copa_bracell.xlsx  (atualiza células SIM/NÃO)
        → dados.json                (alimenta o álbum HTML)

Dependências:
  pip install openpyxl pyodbc python-dotenv

Uso:
  python apuracao_sql.py                   # apuração completa
  python apuracao_sql.py --dry-run         # mostra sem salvar
  python apuracao_sql.py --equipe franca   # só uma equipe
  python apuracao_sql.py --so-manual       # aplica só metas_manuais.json (sem banco)

Metas manuais:
  Edite o arquivo metas_manuais.json antes de rodar.
  Figurinhas marcadas como true substituem o valor do banco.
  Figurinhas ausentes ou false mantêm o valor apurado.
"""

import json
import os
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ══════════════════════════════════════════════════════════════════
#  1b. DATAS DA CAMPANHA
# ══════════════════════════════════════════════════════════════════

DTA_INICIO = "2026-06-01"
DTA_FIM    = "2026-07-31"


# ══════════════════════════════════════════════════════════════════
#  2. DE-PARA: COD_SUPERVISOR → ID EQUIPE
#  Chave  = cod_supervisor exato que vem do banco
#  Valor  = equipe_id que bate com a coluna D da planilha
# ══════════════════════════════════════════════════════════════════

SUPERVISOR_PARA_EQUIPE = {
    "89": "alemanha", #NEIA
    "5": "belgica", #THAYLA
    "249": "canada", #RODRIGO
    "168": "espanha", #RAI
    "1598": "franca", #RONALD
    "1438": "holanda", #ANAIVE
    "1719": "inglaterra", #DEBORA
    "1605": "japao", #EDSON
    "1353": "mexico", #RAMOS
    "1744": "portugal", #PEDRO
    "43": "uruguai", #GISELE
}

# Lista de equipes derivada do de-para (não precisa editar)
EQUIPES = list(SUPERVISOR_PARA_EQUIPE.values())


# ══════════════════════════════════════════════════════════════════
#  3. CAMINHOS DOS ARQUIVOS
# ══════════════════════════════════════════════════════════════════

PASTA_PROJETO        = Path(__file__).parent
ARQUIVO_EXCEL        = PASTA_PROJETO / "controle_copa_bracell.xlsx"
ARQUIVO_JSON    = PASTA_PROJETO / "dados.json"
ARQUIVO_MANUAIS = PASTA_PROJETO / "metas_manuais.json"
ARQUIVO_LOG     = PASTA_PROJETO / "log_apuracao.csv"
ABA_CONTROLE         = "Controle"
LINHA_INICIO         = 3
COL_ID_EQUIPE        = 4   # coluna D
COL_FIGURINHA_INICIO = 5   # coluna E


# ══════════════════════════════════════════════════════════════════
#  4. QUERIES POR FIGURINHA
#  Cada query recebe :cod_supervisor e :periodo e deve retornar
#  exatamente 1 linha com 1 coluna:
#    valor > 0 / True  → meta ATINGIDA → "SIM"
#    valor = 0 / False → não atingida  → ""  (célula vazia)
# ══════════════════════════════════════════════════════════════════

QUERIES_FIGURINHAS = [

    # ── 01 — 90% de Efetividade ───────────────────────────────────
    ("efetividade_90", """
        SELECT CASE WHEN AVG(CAST(efetividade AS FLOAT)) >= 0.90
                    THEN 1 ELSE 0 END
        FROM   apuracao_promotores
        WHERE  cod_supervisor = :cod_supervisor
          AND  periodo        = :periodo
    """),

    # ── 02-03 — Cobertura de Ponto Extra (50% / 100%) ────────────
    # Apurados em lote pela função apurar_pe() — não têm query individual.
    ("pe_50pct_prom",  None),
    ("pe_100pct_prom", None),

    # ── 04-07 — Cobertura por tipo de Ponto Extra ────────────────
    # Apurados em lote pela função apurar_tipos_pe().
    ("cross_50pct",       None),   # CROSS >= 50% das lojas
    ("terminal_20pct",    None),   # PONTA DE GONDOLA >= 20% das lojas
    ("ilha_20pct",        None),   # ILHA >= 20% das lojas
    ("display_turbilhao", None),   # DISPLAY + DISPLAY TURBILHAO >= 5 lojas

    # ── 08 — 2 Lojas PE Criativo Copa ─────────────────────────────
    ("criativo_copa_2", """
        SELECT CASE WHEN COUNT(*) >= 2 THEN 1 ELSE 0 END
        FROM   apuracao_lojas
        WHERE  cod_supervisor = :cod_supervisor
          AND  periodo        = :periodo
          AND  tem_pe_copa    = 1
    """),

    # ── 09 — 4 Lojas PE Criativo Copa ─────────────────────────────
    ("criativo_copa_4", """
        SELECT CASE WHEN COUNT(*) >= 4 THEN 1 ELSE 0 END
        FROM   apuracao_lojas
        WHERE  cod_supervisor = :cod_supervisor
          AND  periodo        = :periodo
          AND  tem_pe_copa    = 1
    """),

    # ── 10 — 2 Lojas PE Criativo São João ─────────────────────────
    ("criativo_sj_2", """
        SELECT CASE WHEN COUNT(*) >= 2 THEN 1 ELSE 0 END
        FROM   apuracao_lojas
        WHERE  cod_supervisor = :cod_supervisor
          AND  periodo        = :periodo
          AND  tem_pe_sjoao   = 1
    """),

    # ── 11 — 4 Lojas PE Criativo São João ─────────────────────────
    ("criativo_sj_4", """
        SELECT CASE WHEN COUNT(*) >= 4 THEN 1 ELSE 0 END
        FROM   apuracao_lojas
        WHERE  cod_supervisor = :cod_supervisor
          AND  periodo        = :periodo
          AND  tem_pe_sjoao   = 1
    """),

    # ── 12-15 — MPDV Leve (25%/50%) e Pesado (1/2 lojas) ─────────
    # Apurados em lote pela função apurar_mpdv() — não têm query
    # individual aqui. Os ids precisam existir na lista para o
    # mapeamento de colunas do Excel funcionar corretamente.
    ("mpdv_leve_25",  None),
    ("mpdv_leve_50",  None),
    ("mpdv_pesado_1", None),
    ("mpdv_pesado_2", None),

    # ── 16 — Anúncio Autofalante ───────────────────────────────────
    ("autofalante", """
        SELECT CASE WHEN COUNT(*) >= 1 THEN 1 ELSE 0 END
        FROM   apuracao_lojas
        WHERE  cod_supervisor = :cod_supervisor
          AND  periodo        = :periodo
          AND  tem_autofalante = 1
    """),
]

IDS_FIGURINHA = [fid for fid, _ in QUERIES_FIGURINHAS]

# Figurinhas com query individual (None = apuradas em lote)
QUERIES_INDIVIDUAIS = [(fid, q) for fid, q in QUERIES_FIGURINHAS if q is not None]


# ══════════════════════════════════════════════════════════════════
#  4b. METAS MANUAIS  (metas_manuais.json)
#
#  Use para figurinhas validadas em campo, sem dados no banco:
#  criativos Copa/São João (08–11), autofalante (16), etc.
#
#  Formato do arquivo:
#  {
#    "franca": {
#      "criativo_copa_2":  true,
#      "criativo_copa_4":  false,
#      "criativo_sj_2":    true,
#      "criativo_sj_4":    false,
#      "autofalante":      true
#    },
#    "alemanha": { ... }
#  }
#
#  Regras:
#    - true  → marca SIM (sobrescreve o banco se houver conflito)
#    - false → remove SIM (sobrescreve o banco se houver conflito)
#    - chave ausente → mantém o valor apurado pelo banco
#
#  Para marcar APENAS uma equipe, coloque só ela no arquivo.
#  As demais equipes ficam intactas.
# ══════════════════════════════════════════════════════════════════

FIGURINHAS_MANUAIS = [
    "efetividade_90",    # fig 01 — caso não haja query de efetividade
    "criativo_copa_2",   # fig 08
    "criativo_copa_4",   # fig 09
    "criativo_sj_2",     # fig 10
    "criativo_sj_4",     # fig 11
    "autofalante",       # fig 16
]

def carregar_metas_manuais() -> dict:
    """
    Lê metas_manuais.json e retorna { equipe_id: { fig_id: bool } }.
    Retorna dict vazio se o arquivo não existir (não é erro).
    """
    if not ARQUIVO_MANUAIS.exists():
        return {}
    try:
        raw = json.loads(ARQUIVO_MANUAIS.read_text(encoding="utf-8"))
        # Normaliza chaves para lowercase
        return {
            eq.lower(): {fig: bool(v) for fig, v in figs.items()}
            for eq, figs in raw.items()
        }
    except Exception as e:
        print(f"⚠️  Erro ao ler metas_manuais.json: {e}")
        return {}


def mesclar_manuais(resultados: dict, manuais: dict) -> dict:
    """
    Aplica as metas manuais sobre os resultados do banco.
    Só sobrescreve as figurinhas presentes no arquivo manual.
    """
    if not manuais:
        return resultados

    for equipe_id, figs in manuais.items():
        if equipe_id not in resultados:
            # Equipe existe só no manual — cria entrada zerada e aplica
            resultados[equipe_id] = {fid: False for fid in IDS_FIGURINHA}
        for fig_id, valor in figs.items():
            if fig_id in IDS_FIGURINHA:
                resultados[equipe_id][fig_id] = valor
                status = "✅ manual" if valor else "· manual (removido)"
                print(f"  {equipe_id:12s} | {fig_id:25s} → {status}")
            else:
                print(f"  ⚠️  ID desconhecido em metas_manuais.json: '{fig_id}' — ignorado.")
    return resultados


# ══════════════════════════════════════════════════════════════════
#  5. QUERY MPDV EM LOTE  (figurinhas 12-15)
#  Retorna todas as equipes de uma vez, evitando N chamadas ao banco.
#  O resultado e indexado por COD_SUPERVISOR para lookup O(1).
# ══════════════════════════════════════════════════════════════════

QUERY_MPDV = """
WITH
  LOJAS_ABSOLUTO AS (
    SELECT A.COD_LOJA, A.COD_SKU,
           SUM(A.FL_PRESENTE)   AS QTDE_PRESENCA,
           COUNT(A.FL_PRESENTE) AS QTDE_PESQUISA,
           CAST(SUM(A.FL_PRESENTE) * 1.0 / COUNT(A.FL_PRESENTE) AS DECIMAL(10,2)) AS PRESENCA
    FROM   bi_f_produto  AS A
    INNER JOIN bi_d_sku  AS B ON A.COD_MARCA = B.COD_MARCA
    INNER JOIN bi_d_loja AS C ON A.COD_LOJA  = C.COD_LOJA
    WHERE  B.DES_MARCA  = 'ABSOLUTO'
      AND  A.DATA       BETWEEN ? AND ?
      AND  C.DES_REGIAO IN ('SUL', 'SUDESTE', 'CENTRO-OESTE')
    GROUP BY A.COD_LOJA, A.COD_SKU
    HAVING CAST(SUM(A.FL_PRESENTE) * 1.0 / COUNT(A.FL_PRESENTE) AS DECIMAL(10,2)) >= 0.2
  ),
  BASE_LOJAS AS (
    SELECT A.COD_SUPERVISOR,
           SUPERVISOR,
           COUNT(DISTINCT D.COD_LOJA) AS QTDE_LOJAS
    FROM   bi_f_efetividade AS A
    INNER JOIN bi_d_pessoa  AS B ON A.COD_SUPERVISOR = B.COD_SUPERVISOR
    INNER JOIN bi_d_loja    AS C ON A.COD_LOJA       = C.COD_LOJA
    INNER JOIN LOJAS_ABSOLUTO AS D ON D.COD_LOJA     = A.COD_LOJA
    WHERE  A.DATA       BETWEEN ? AND ?
      AND  C.DES_REGIAO IN ('SUL', 'SUDESTE', 'CENTRO-OESTE')
    GROUP BY A.COD_SUPERVISOR, SUPERVISOR
  )
SELECT
    A.COD_SUPERVISOR,
    C.QTDE_LOJAS,
    COUNT(DISTINCT A.COD_LOJA)                                                          AS QTDE_LOJAS_MPDV,
    ROUND(CAST(COUNT(DISTINCT A.COD_LOJA) AS FLOAT) / NULLIF(C.QTDE_LOJAS, 0) * 100, 2) AS PERC_MPDV,
    B.FL_PESADO,
    B.FL_LEVE
FROM   bi_f_mpdv      AS A
INNER JOIN bi_d_mpdv  AS B ON A.COD_MPDV      = B.COD_MPDV
INNER JOIN BASE_LOJAS AS C ON A.COD_SUPERVISOR = C.COD_SUPERVISOR
WHERE  A.DATA       BETWEEN ? AND ?
  AND  A.FL_PRESENTE = '1'
GROUP BY A.COD_SUPERVISOR, C.QTDE_LOJAS, B.FL_PESADO, B.FL_LEVE
"""

def apurar_mpdv(cursor) -> dict:
    """
    Executa a query MPDV em lote e devolve:
    {
      "COD001": {
        "mpdv_leve_25":  True/False,
        "mpdv_leve_50":  True/False,
        "mpdv_pesado_1": True/False,
        "mpdv_pesado_2": True/False,
      },
      ...
    }
    """

    print(f"  📦 MPDV em lote  [{DTA_INICIO} → {DTA_FIM}]")
    try:
        cursor.execute(QUERY_MPDV, (DTA_INICIO, DTA_FIM) * 3)
        rows = cursor.fetchall()
        print(f"  🔍 MPDV: {len(rows)} linha(s) retornada(s)")
        if rows:
            print(f"     Amostra linha 1: {rows[0]}")
    except Exception as e:
        print(f"  ⚠️  Erro na query MPDV: {e}")
        return {}

    # Agrupa por COD_SUPERVISOR; uma linha FL_LEVE e outra FL_PESADO
    agrupado: dict[str, dict] = {}
    for cod_sup, qtde_lojas, qtde_mpdv, perc_mpdv, fl_pesado, fl_leve in rows:
        cod_sup = str(cod_sup).strip()
        if cod_sup not in agrupado:
            agrupado[cod_sup] = {
                "mpdv_leve_25":  False,
                "mpdv_leve_50":  False,
                "mpdv_pesado_1": False,
                "mpdv_pesado_2": False,
            }
        if fl_leve == 1 or fl_leve == "1":
            agrupado[cod_sup]["mpdv_leve_25"] = float(perc_mpdv or 0) >= 25.0
            agrupado[cod_sup]["mpdv_leve_50"] = float(perc_mpdv or 0) >= 50.0
        if fl_pesado == 1 or fl_pesado == "1":
            agrupado[cod_sup]["mpdv_pesado_1"] = int(qtde_mpdv or 0) >= 1
            agrupado[cod_sup]["mpdv_pesado_2"] = int(qtde_mpdv or 0) >= 2

    return agrupado


# ══════════════════════════════════════════════════════════════════
#  5b. QUERY TIPOS DE PONTO EXTRA EM LOTE  (figurinhas 04–07)
#  Retorna uma linha por COD_SUPERVISOR + DES_TIPO_PONTO_EXTRA:
#    CROSS            → fig04  cross_50pct       PERC_LOJAS >= 50%
#    PONTA DE GONDOLA → fig05  terminal_20pct    PERC_LOJAS >= 20%
#    ILHA             → fig06  ilha_20pct        PERC_LOJAS >= 20%
#    DISPLAY          → fig07  display_turbilhao QTDE_LOJAS_PE >= 5
#    DISPLAY TURBILHAO→ fig07  display_turbilhao (acumula com DISPLAY)
# ══════════════════════════════════════════════════════════════════

QUERY_TIPOS_PE = """
WITH
  LOJAS_ABSOLUTO AS (
    SELECT A.COD_LOJA, A.COD_SKU,
           SUM(A.FL_PRESENTE)   AS QTDE_PRESENCA,
           COUNT(A.FL_PRESENTE) AS QTDE_PESQUISA,
           CAST(SUM(A.FL_PRESENTE) * 1.0 / COUNT(A.FL_PRESENTE) AS DECIMAL(10,2)) AS PRESENCA
    FROM   bi_f_produto  AS A
    INNER JOIN bi_d_sku  AS B ON A.COD_MARCA = B.COD_MARCA
    INNER JOIN bi_d_loja AS C ON A.COD_LOJA  = C.COD_LOJA
    WHERE  B.DES_MARCA  = 'ABSOLUTO'
      AND  A.DATA       BETWEEN ? AND ?
      AND  C.DES_REGIAO IN ('SUL', 'SUDESTE', 'CENTRO-OESTE')
    GROUP BY A.COD_LOJA, A.COD_SKU
    HAVING CAST(SUM(A.FL_PRESENTE) * 1.0 / COUNT(A.FL_PRESENTE) AS DECIMAL(10,2)) >= 0.2
  ),
  BASE_LOJAS AS (
    SELECT A.COD_SUPERVISOR,
           SUPERVISOR,
           COUNT(DISTINCT D.COD_LOJA) AS QTDE_LOJAS
    FROM   bi_f_efetividade AS A
    INNER JOIN bi_d_pessoa  AS B ON A.COD_SUPERVISOR = B.COD_SUPERVISOR
    INNER JOIN bi_d_loja    AS C ON A.COD_LOJA       = C.COD_LOJA
    INNER JOIN LOJAS_ABSOLUTO AS D ON D.COD_LOJA     = A.COD_LOJA
    WHERE  A.DATA       BETWEEN ? AND ?
      AND  C.DES_REGIAO IN ('SUL', 'SUDESTE', 'CENTRO-OESTE')
    GROUP BY A.COD_SUPERVISOR, SUPERVISOR
  )
SELECT
    A.COD_SUPERVISOR,
    C.QTDE_LOJAS,
    COUNT(DISTINCT A.COD_LOJA)                                                          AS QTDE_LOJAS_PE,
    ROUND(CAST(COUNT(DISTINCT A.COD_LOJA) AS FLOAT) / NULLIF(C.QTDE_LOJAS, 0) * 100, 2) AS PERC_LOJAS,
    B.DES_TIPO_PONTO_EXTRA
FROM   bi_f_ponto_extra          AS A
INNER JOIN bi_d_tipo_ponto_extra AS B ON A.COD_TIPO_PONTO_EXTRA = B.COD_TIPO_PONTO_EXTRA
INNER JOIN BASE_LOJAS            AS C ON A.COD_SUPERVISOR       = C.COD_SUPERVISOR
WHERE  A.DATA               BETWEEN ? AND ?
  AND  B.DES_TIPO_PONTO_EXTRA IN ('ILHA', 'PONTA DE GONDOLA', 'CROSS', 'DISPLAY', 'DISPLAY TURBILHAO')
  AND  A.FL_PE_EXISTE = '1'
GROUP BY A.COD_SUPERVISOR, C.QTDE_LOJAS, B.DES_TIPO_PONTO_EXTRA
"""

def apurar_tipos_pe(cursor) -> dict:
    """
    Executa a query de tipos de PE em lote e devolve:
    {
      "COD001": {
        "cross_50pct":       True/False,   # CROSS >= 50%
        "terminal_20pct":    True/False,   # PONTA DE GONDOLA >= 20%
        "ilha_20pct":        True/False,   # ILHA >= 20%
        "display_turbilhao": True/False,   # DISPLAY + DISPLAY TURBILHAO >= 5 lojas
      },
      ...
    }
    DISPLAY e DISPLAY TURBILHAO são somados: se juntos >= 5 lojas únicas = atingiu.
    """

    print(f"  📦 Tipos de PE em lote  [{DTA_INICIO} → {DTA_FIM}]")
    try:
        cursor.execute(QUERY_TIPOS_PE, (DTA_INICIO, DTA_FIM) * 3)
        rows = cursor.fetchall()
        print(f"  🔍 Tipos PE: {len(rows)} linha(s) retornada(s)")
        if rows:
            print(f"     Amostra linha 1: {rows[0]}")
    except Exception as e:
        print(f"  ⚠️  Erro na query Tipos PE: {e}")
        return {}

    # Acumula por supervisor; DISPLAY e DISPLAY TURBILHAO somam para fig07
    # Cada linha: COD_SUPERVISOR, QTDE_LOJAS, QTDE_LOJAS_PE, PERC_LOJAS, DES_TIPO_PONTO_EXTRA
    acumulado: dict = {}
    for cod_sup, qtde_lojas, qtde_pe, perc, tipo in rows:
        cod_sup = str(cod_sup).strip()
        tipo    = str(tipo).strip().upper()
        perc    = float(perc   or 0)
        qtde_pe = int(qtde_pe  or 0)

        if cod_sup not in acumulado:
            acumulado[cod_sup] = {
                "cross_50pct":       False,
                "terminal_20pct":    False,
                "ilha_20pct":        False,
                "display_qtde":      0,      # acumulador interno
                "display_turbilhao": False,
            }

        if tipo == "CROSS":
            acumulado[cod_sup]["cross_50pct"]    = perc >= 50.0
        elif tipo == "PONTA DE GONDOLA":
            acumulado[cod_sup]["terminal_20pct"] = perc >= 20.0
        elif tipo == "ILHA":
            acumulado[cod_sup]["ilha_20pct"]     = perc >= 20.0
        elif tipo in ("DISPLAY", "DISPLAY TURBILHAO"):
            acumulado[cod_sup]["display_qtde"]  += qtde_pe

    # Resolve flag display após acumular ambos os tipos
    resultado = {}
    for cod_sup, vals in acumulado.items():
        resultado[cod_sup] = {
            "cross_50pct":       vals["cross_50pct"],
            "terminal_20pct":    vals["terminal_20pct"],
            "ilha_20pct":        vals["ilha_20pct"],
            "display_turbilhao": vals["display_qtde"] >= 5,
        }

    return resultado
#  Retorna COBERTURA (0.00–1.00) por COD_SUPERVISOR:
#    >= 0.50  → fig 02  pe_50pct_prom
#    >= 1.00  → fig 03  pe_100pct_prom
# ══════════════════════════════════════════════════════════════════

QUERY_PE = """
WITH
  LOJAS_ABSOLUTO AS (
    SELECT A.COD_LOJA, A.COD_SKU,
           SUM(A.FL_PRESENTE)   AS QTDE_PRESENCA,
           COUNT(A.FL_PRESENTE) AS QTDE_PESQUISA,
           CAST(SUM(A.FL_PRESENTE) * 1.0 / COUNT(A.FL_PRESENTE) AS DECIMAL(10,2)) AS PRESENCA
    FROM   bi_f_produto  AS A
    INNER JOIN bi_d_sku  AS B ON A.COD_MARCA = B.COD_MARCA
    INNER JOIN bi_d_loja AS C ON A.COD_LOJA  = C.COD_LOJA
    WHERE  B.DES_MARCA  = 'ABSOLUTO'
      AND  A.DATA       BETWEEN ? AND ?
      AND  C.DES_REGIAO IN ('SUL', 'SUDESTE', 'CENTRO-OESTE')
    GROUP BY A.COD_LOJA, A.COD_SKU
    HAVING CAST(SUM(A.FL_PRESENTE) * 1.0 / COUNT(A.FL_PRESENTE) AS DECIMAL(10,2)) >= 0.2
  ),
  BASE_ROTAS AS (
    SELECT MAX(DTA_ROTEIRO) AS DATA, COD_ROTA
    FROM   bi_d_hierarquia_rota
    WHERE  DTA_ROTEIRO BETWEEN ? AND ?
    GROUP BY COD_ROTA
  ),
  BASE_EQUIPES AS (
    SELECT A.COD_ROTA, A.COD_SUPERVISOR
    FROM   bi_f_efetividade   AS A
    INNER JOIN BASE_ROTAS     AS B ON A.DATA     = B.DATA AND A.COD_ROTA = B.COD_ROTA
    INNER JOIN LOJAS_ABSOLUTO AS C ON C.COD_LOJA = A.COD_LOJA
    WHERE  A.DATA BETWEEN ? AND ?
    GROUP BY A.COD_ROTA, A.COD_SUPERVISOR
  ),
  QTDE_VAGAS AS (
    SELECT COD_SUPERVISOR,
           COUNT(DISTINCT COD_ROTA) AS QTDE_VAGAS
    FROM   BASE_EQUIPES
    GROUP BY COD_SUPERVISOR
  )
SELECT
    C.COD_SUPERVISOR,
    COUNT(DISTINCT A.COD_ROTA)                                                           AS QTDE_PROM_PE,
    E.QTDE_VAGAS,
    CAST(COUNT(DISTINCT A.COD_ROTA) * 1.0 / NULLIF(E.QTDE_VAGAS, 0) AS DECIMAL(10,2))  AS COBERTURA
FROM   [bi_s3_bracell].[dbo].[bi_f_ponto_extra] AS A
INNER JOIN bi_d_marca   AS B ON A.COD_MARCA      = B.COD_MARCA
INNER JOIN BASE_EQUIPES AS C ON A.COD_ROTA       = C.COD_ROTA
INNER JOIN bi_d_pessoa  AS D ON A.COD_PESSOA     = D.COD_PESSOA
INNER JOIN QTDE_VAGAS   AS E ON C.COD_SUPERVISOR = E.COD_SUPERVISOR
WHERE  B.DES_MARCA    = 'ABSOLUTO'
  AND  A.FL_PE_EXISTE = 1
  AND  A.DATA         BETWEEN ? AND ?
  AND  D.PERFIL       IN ('FERISTA', 'PROMOTOR')
GROUP BY C.COD_SUPERVISOR, E.QTDE_VAGAS
"""

def apurar_pe(cursor) -> dict:
    """
    Executa a query de Ponto Extra em lote e devolve:
    {
      "COD001": {"pe_50pct_prom": True/False, "pe_100pct_prom": True/False},
      ...
    }
    """

    print(f"  📦 Ponto Extra em lote  [{DTA_INICIO} → {DTA_FIM}]")
    try:
        cursor.execute(QUERY_PE, (DTA_INICIO, DTA_FIM) * 4)
        rows = cursor.fetchall()
        print(f"  🔍 PE: {len(rows)} linha(s) retornada(s)")
        if rows:
            print(f"     Amostra linha 1: {rows[0]}")
    except Exception as e:
        print(f"  ⚠️  Erro na query Ponto Extra: {e}")
        return {}

    # Cada linha: COD_SUPERVISOR, QTDE_PROM_PE, QTDE_VAGAS, COBERTURA
    resultado = {}
    for cod_sup, qtde_prom_pe, qtde_vagas, cobertura in rows:
        cod_sup   = str(cod_sup).strip()
        cobertura = float(cobertura or 0)
        resultado[cod_sup] = {
            "pe_50pct_prom":  cobertura >= 0.50,
            "pe_100pct_prom": cobertura >= 1.00,
        }

    return resultado


# ══════════════════════════════════════════════════════════════════
#  6. CONEXÃO
# ══════════════════════════════════════════════════════════════════

def criar_conexao():
    import pyodbc
    dados_conexao = ( 
    'DRIVER={SQL Server};'
    'SERVER=172.18.0.59;'
    'DATABASE=bi_s3_bracell;'
)
    return pyodbc.connect(dados_conexao)


def executar_query(cursor, sql, cod_supervisor):
    try:
        cursor.execute(sql, (cod_supervisor,))
        row = cursor.fetchone()
        if row is None:
            return False
        v = row[0]
        return bool(v) and v not in (0, "0", False, None)
    except Exception as e:
        print(f"  ⚠️  Erro na query ({cod_supervisor}): {e}")
        return False


# ══════════════════════════════════════════════════════════════════
#  7. APURAÇÃO VIA SQL
# ══════════════════════════════════════════════════════════════════

def apurar_sql(equipes_filtro=None):
    """Consulta o banco e retorna { equipe_id: { fig_id: bool } }."""
    equipes_alvo    = set(equipes_filtro or EQUIPES)
    equipe_para_cod = {v: k for k, v in SUPERVISOR_PARA_EQUIPE.items()}

    print(f"\n🔌 Conectando ao SQL Server…")
    try:
        conn = criar_conexao()
    except Exception as e:
        print(f"❌ Falha na conexão: {e}")
        print("   Verifique DB_SERVER e DB_NAME no topo do script.")
        sys.exit(1)

    cursor = conn.cursor()
    print(f"✅ Conectado!\n")

    # ── 1. Apura em lote: MPDV, Tipos de PE e Cobertura PE ───────
    print("🔎 Apurando MPDV (lote)…")
    mpdv_por_cod = apurar_mpdv(cursor)

    print("🔎 Apurando Tipos de Ponto Extra (lote)…")
    tipos_pe_por_cod = apurar_tipos_pe(cursor)

    print("🔎 Apurando Cobertura de Promotores c/ PE (lote)…")
    pe_por_cod = apurar_pe(cursor)

    # ── 2. Apura demais figurinhas individualmente por equipe ──────
    resultados = {}
    for equipe_id in equipes_alvo:
        cod = equipe_para_cod.get(equipe_id)
        if not cod:
            print(f"⚠️  Equipe '{equipe_id}' sem cod_supervisor no de-para — pulando.")
            continue

        print(f"\n🏳  {equipe_id}  (cod_supervisor: {cod})")
        resultados[equipe_id] = {}

        for fig_id, query in QUERIES_INDIVIDUAIS:
            atingiu = executar_query(cursor, query, cod)
            resultados[equipe_id][fig_id] = atingiu
            print(f"   {'✅' if atingiu else '·'} {fig_id}")

        # Mescla Cobertura Promotores c/ PE (figs 02–03)
        pe_eq = pe_por_cod.get(cod, {})
        for fig in ("pe_50pct_prom", "pe_100pct_prom"):
            atingiu = pe_eq.get(fig, False)
            resultados[equipe_id][fig] = atingiu
            print(f"   {'✅' if atingiu else '·'} {fig}")

        # Mescla Tipos de PE (figs 04–07)
        tpe_eq = tipos_pe_por_cod.get(cod, {})
        for fig in ("cross_50pct", "terminal_20pct", "ilha_20pct", "display_turbilhao"):
            atingiu = tpe_eq.get(fig, False)
            resultados[equipe_id][fig] = atingiu
            print(f"   {'✅' if atingiu else '·'} {fig}")

        # Mescla MPDV (figs 12–15)
        mpdv_eq = mpdv_por_cod.get(cod, {})
        for fig in ("mpdv_leve_25", "mpdv_leve_50", "mpdv_pesado_1", "mpdv_pesado_2"):
            atingiu = mpdv_eq.get(fig, False)
            resultados[equipe_id][fig] = atingiu
            print(f"   {'✅' if atingiu else '·'} {fig}")

    cursor.close()
    conn.close()
    print("\n🔒 Conexão encerrada.")
    return resultados


# ══════════════════════════════════════════════════════════════════
#  7. ATUALIZAÇÃO DO EXCEL
# ══════════════════════════════════════════════════════════════════

def _carregar_log() -> dict:
    """
    Lê o log existente e retorna o estado ATUAL de cada (equipe, figurinha).
    Estado atual = último evento registrado para aquela chave.
    Retorna { (equipe, figurinha): ultimo_status }
    onde ultimo_status é 'preenchida', 'mantida', 'removida' ou 'nao_atingida'.
    Compatível com logs gerados pela versão anterior (só tinham preenchida/removida).
    """
    import csv
    estado = {}   # (equipe, figurinha) → ultimo_status
    if not ARQUIVO_LOG.exists():
        return estado
    with open(ARQUIVO_LOG, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        # Normaliza nomes de colunas (remove BOM residual e espaços)
        if reader.fieldnames:
            reader.fieldnames = [c.strip().lstrip("\ufeff") for c in reader.fieldnames]
        for i, row in enumerate(reader):
            try:
                acao   = (row.get("acao") or "").strip()
                equipe = (row.get("equipe") or "").strip()
                fig    = (row.get("figurinha") or "").strip()
                if not equipe or not fig:
                    continue
                # Linhas antigas sem status 'mantida'/'nao_atingida' são aceitas
                if acao not in ("preenchida", "mantida", "removida", "nao_atingida"):
                    continue
                estado[(equipe, fig)] = acao
            except Exception as e:
                print(f"  ⚠️  Log linha {i+2} ignorada: {e} | {dict(row)}")
    return estado


def _gravar_log(novos_eventos: list[dict]):
    """Acrescenta todos os eventos da execução ao CSV (nunca sobrescreve)."""
    import csv
    novo_arquivo = not ARQUIVO_LOG.exists()
    with open(ARQUIVO_LOG, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f, fieldnames=["data_hora", "equipe", "figurinha", "acao", "origem"]
        )
        if novo_arquivo:
            writer.writeheader()
        for ev in novos_eventos:
            writer.writerow(ev)


def ranking_album() -> list[dict]:
    """
    Lê o log e calcula quem completou o álbum e em que data/hora.
    Data de conclusão = max( data_inicio_posse_continua de cada figurinha ).
    data_inicio_posse_continua = data do primeiro 'preenchida' ou 'mantida'
    da última sequência ininterrupta (sem 'removida' depois).
    Compatível com logs gerados pela versão anterior.
    """
    import csv
    from collections import defaultdict

    if not ARQUIVO_LOG.exists():
        return []

    # Agrupa todos os eventos por (equipe, figurinha), em ordem de leitura
    historico: dict = defaultdict(list)
    with open(ARQUIVO_LOG, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [c.strip().lstrip("\ufeff") for c in reader.fieldnames]
        for i, row in enumerate(reader):
            try:
                equipe = (row.get("equipe") or "").strip()
                fig    = (row.get("figurinha") or "").strip()
                if not equipe or not fig:
                    continue
                historico[(equipe, fig)].append({
                    "acao":      (row.get("acao") or "").strip(),
                    "data_hora": (row.get("data_hora") or "").strip(),
                })
            except Exception as e:
                print(f"  ⚠️  Ranking linha {i+2} ignorada: {e}")

    # Para cada (equipe, figurinha) determina a data mínima da última posse contínua:
    # percorre de trás para frente até encontrar uma 'removida' ou o início do histórico.
    posse_min: dict = defaultdict(dict)  # equipe → {fig → data_hora}
    for (equipe, fig), eventos in historico.items():
        data_inicio_posse = None
        for ev in reversed(eventos):
            if ev["acao"] in ("preenchida", "mantida"):
                data_inicio_posse = ev["data_hora"]
            elif ev["acao"] == "removida":
                break   # sequência contínua interrompida
        if data_inicio_posse:
            posse_min[equipe][fig] = data_inicio_posse

    n = len(IDS_FIGURINHA)
    todas_equipes = set(eq for (eq, _) in historico.keys())
    ranking = []
    for equipe in todas_equipes:
        figs_com_posse = posse_min.get(equipe, {})
        total          = len(figs_com_posse)
        if total == n:
            data_conclusao = max(figs_com_posse.values())
            ranking.append({"equipe": equipe, "total": total, "concluido_em": data_conclusao})
        else:
            ranking.append({"equipe": equipe, "total": total, "concluido_em": None})

    ranking.sort(key=lambda x: (x["concluido_em"] is None, x["concluido_em"], -x["total"]))
    return ranking


def atualizar_excel(resultados: dict, dry_run: bool = False, origem: str = "banco"):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    print(f"\n📊 {'[DRY-RUN] ' if dry_run else ''}Atualizando: {ARQUIVO_EXCEL.name}")

    wb = load_workbook(ARQUIVO_EXCEL)
    ws = wb[ABA_CONTROLE]

    VERDE     = PatternFill("solid", fgColor="C6EFCE")
    BRANCO    = PatternFill("solid", fgColor="FFFFFF")
    FONTE_SIM = Font(bold=True,  color="276221")
    FONTE_NAO = Font(bold=False, color="999999")

    # Estado atual do log — último status registrado por (equipe, figurinha)
    estado_log = _carregar_log()
    agora      = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    eventos    = []   # todos os eventos desta execução (uma linha por figurinha)
    alteracoes = 0

    for linha in ws.iter_rows(min_row=LINHA_INICIO):
        equipe_id = str(linha[COL_ID_EQUIPE - 1].value or "").strip().lower()
        if equipe_id not in resultados:
            continue

        eq_res       = resultados[equipe_id]
        conquistadas = 0

        for i, fig_id in enumerate(IDS_FIGURINHA):
            cell    = linha[COL_FIGURINHA_INICIO + i - 1]
            atingiu = eq_res.get(fig_id, False)
            novo    = "SIM" if atingiu else ""
            atual   = str(cell.value or "").strip().upper()
            chave   = (equipe_id, fig_id)
            ultimo  = estado_log.get(chave)   # status da última execução

            if novo != atual:
                alteracoes += 1
                print(f"  {equipe_id:12s} | #{i+1:02d} {fig_id:25s} → {'SIM ✅' if atingiu else 'removido'}")

            # ── Determina o status desta figurinha nesta execução ──────────
            if atingiu and ultimo not in ("preenchida", "mantida"):
                acao = "preenchida"      # conquistada agora (era removida/nao_atingida/nova)
            elif atingiu:
                acao = "mantida"         # já tinha, continua tendo
            elif not atingiu and ultimo in ("preenchida", "mantida"):
                acao = "removida"        # tinha, perdeu
            else:
                acao = "nao_atingida"    # nunca teve (ou já estava removida)

            eventos.append({
                "data_hora": agora,
                "equipe":    equipe_id,
                "figurinha": fig_id,
                "acao":      acao,
                "origem":    origem,
            })

            if not dry_run:
                cell.value     = novo
                cell.fill      = VERDE if atingiu else BRANCO
                cell.font      = FONTE_SIM if atingiu else FONTE_NAO
                cell.alignment = Alignment(horizontal="center")

            if atingiu:
                conquistadas += 1

        col_prog = COL_FIGURINHA_INICIO + len(IDS_FIGURINHA) - 1
        if not dry_run:
            linha[col_prog].value = f"{conquistadas}/{len(IDS_FIGURINHA)}"

    if dry_run:
        print(f"\n  {alteracoes} alteração(ões) seriam feitas. (dry-run — nada salvo)")
        return

    print(f"\n  ✅ {alteracoes} alteração(ões) realizada(s).")
    try:
        ws["A2"] = (f"COPA BRACELL ABSOLUTO — CONTROLE DE FIGURINHAS 2026  "
                    f"| Última atualização: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    except Exception:
        pass
    wb.save(ARQUIVO_EXCEL)
    print(f"  💾 Salvo: {ARQUIVO_EXCEL}")

    _gravar_log(eventos)
    qtd_novas     = sum(1 for e in eventos if e["acao"] == "preenchida")
    qtd_mantidas  = sum(1 for e in eventos if e["acao"] == "mantida")
    qtd_removidas = sum(1 for e in eventos if e["acao"] == "removida")
    print(f"  📋 Log: +{len(eventos)} evento(s) "
          f"[✅ {qtd_novas} novas | 🔒 {qtd_mantidas} mantidas "
          f"| ❌ {qtd_removidas} removidas] → {ARQUIVO_LOG.name}")

    if qtd_novas > 0 or qtd_removidas > 0:
        _exibir_ranking()


# ══════════════════════════════════════════════════════════════════
#  8. GERAÇÃO DO dados.json
# ══════════════════════════════════════════════════════════════════

def _exibir_ranking():
    """Imprime o ranking de conclusão do álbum no terminal."""
    ranking = ranking_album()
    if not ranking:
        return
    n = len(IDS_FIGURINHA)
    print("\n" + "═" * 60)
    print("🏆  RANKING — CONCLUSÃO DO ÁLBUM")
    print("═" * 60)
    pos = 1
    for r in ranking:
        if r["concluido_em"]:
            print(f"  {pos}º  {r['equipe']:15s}  ✅ Completo em {r['concluido_em']}")
            pos += 1
        else:
            print(f"  -   {r['equipe']:15s}  {r['total']:2d}/{n} figurinhas")
    print("═" * 60)


def gerar_json(resultados: dict):
    dados = {"equipes": {}, "gerado_em": datetime.now().isoformat()}
    for eq, figs in resultados.items():
        dados["equipes"][eq] = {fig: True for fig, v in figs.items() if v}
    ARQUIVO_JSON.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  🎴 dados.json gerado: {ARQUIVO_JSON}")





def main():
    parser = argparse.ArgumentParser(description="Copa Bracell — Apuração SQL Server")
    parser.add_argument("--equipe",    help="Filtra para uma única equipe (ex: franca)")
    parser.add_argument("--dry-run",   action="store_true", help="Mostra as mudanças sem salvar")
    parser.add_argument("--so-manual", action="store_true",
                        help="Aplica só metas_manuais.json (sem conectar ao banco)")
    args = parser.parse_args()

    equipes_filtro = [args.equipe.lower()] if args.equipe else None

    # ── Carrega metas manuais (sempre, independente do modo) ──────
    print("📋 Carregando metas_manuais.json…")
    manuais = carregar_metas_manuais()
    if manuais:
        equipes_manual = ", ".join(sorted(manuais.keys()))
        total_flags    = sum(len(v) for v in manuais.values())
        print(f"   {total_flags} meta(s) manual(is) para: {equipes_manual}")
    else:
        print("   (arquivo não encontrado ou vazio — nenhuma meta manual aplicada)")

    # ── Apuração via banco ou só manual ───────────────────────────
    if args.so_manual:
        print("\n⚡ Modo --so-manual: pulando conexão ao banco.")
        # Parte do estado atual do Excel para não zerar o que já foi apurado
        resultados = _ler_estado_excel(equipes_filtro)
        print(f"\n📝 Aplicando metas manuais…")
        resultados = mesclar_manuais(resultados, _filtrar_manuais(manuais, equipes_filtro))
    else:
        resultados = apurar_sql(equipes_filtro)
        if manuais:
            print(f"\n📝 Aplicando metas manuais…")
            resultados = mesclar_manuais(resultados, _filtrar_manuais(manuais, equipes_filtro))

    if args.so_manual:
        print("\n⚡ Modo --so-manual: pulando conexão ao banco.")
        resultados = _ler_estado_excel(equipes_filtro)
        print(f"\n📝 Aplicando metas manuais…")
        resultados = mesclar_manuais(resultados, _filtrar_manuais(manuais, equipes_filtro))
        origem = "manual"
    else:
        resultados = apurar_sql(equipes_filtro)
        if manuais:
            print(f"\n📝 Aplicando metas manuais…")
            resultados = mesclar_manuais(resultados, _filtrar_manuais(manuais, equipes_filtro))
        origem = "banco"

    atualizar_excel(resultados, dry_run=args.dry_run, origem=origem)

    if not args.dry_run:
        print("\n📄 Gerando dados.json…")
        gerar_json(resultados)

    print("\n" + "═" * 60)
    print("RESUMO DA APURAÇÃO")
    print("═" * 60)
    for eq, figs in resultados.items():
        total = sum(1 for v in figs.values() if v)
        n     = len(IDS_FIGURINHA)
        print(f"  {eq:15s}  {total:2d}/{n}  {'█' * total}{'░' * (n - total)}")
    print("═" * 60)
    print(f"Concluído: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")


def _filtrar_manuais(manuais: dict, equipes_filtro) -> dict:
    """Restringe manuais às equipes do filtro (se houver)."""
    if not equipes_filtro:
        return manuais
    return {eq: v for eq, v in manuais.items() if eq in equipes_filtro}


def _ler_estado_excel(equipes_filtro) -> dict:
    """
    Lê o estado atual da planilha para o modo --so-manual,
    evitando zerar figurinhas já apuradas pelo banco.
    """
    from openpyxl import load_workbook
    wb = load_workbook(ARQUIVO_EXCEL)
    ws = wb[ABA_CONTROLE]
    equipes_alvo = set(equipes_filtro or EQUIPES)
    resultados   = {}
    for linha in ws.iter_rows(min_row=LINHA_INICIO):
        equipe_id = str(linha[COL_ID_EQUIPE - 1].value or "").strip().lower()
        if equipe_id not in equipes_alvo:
            continue
        resultados[equipe_id] = {}
        for i, fig_id in enumerate(IDS_FIGURINHA):
            cell  = linha[COL_FIGURINHA_INICIO + i - 1]
            valor = str(cell.value or "").strip().upper()
            resultados[equipe_id][fig_id] = (valor == "SIM")
    return resultados


if __name__ == "__main__":
    main()